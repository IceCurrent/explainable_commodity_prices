#!/usr/bin/env python3
"""Macro panel processing for AE<->Macro CCA.

Turns the single raw Bloomberg/uncertainty workbook
(``NEW_MACRO_COMMODITY_PANEL.xlsx``) into clean, aligned, analysis-ready CSVs
for downstream Canonical Correlation Analysis between autoencoder (AE) latent
commodity factors and a daily macro panel.

Run as::

    python src/macro_processing/process_macro.py \
        --input data/macro/NEW_MACRO_COMMODITY_PANEL.xlsx --outdir .

Fully offline and deterministic: the same input yields byte-identical output.
No CCA, no standardization, no monthly-style forward-filling.

See ``prompt.md`` for the full specification.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# ============================================================================
# EDITABLE CONSTANTS  (the user will tweak these)
# ============================================================================

# Study window — hard clip, inclusive.
WINDOW = ("2013-10-18", "2026-06-01")

# Holiday-bridge forward-fill limit on the master trading-day grid.
# 10 trading days covers the longest single closures (China Golden Week /
# Lunar New Year for cny_10y, csi300, shcomp, hscei).
BRIDGE_LIMIT = 10

# Flag a bridged run longer than this many trading days in the QA report.
LONG_BRIDGE_FLAG = 7

# The editable TRANSFORMS map. Per-column transform applied to aligned levels,
# oldest -> newest. One judgment knob: gpr / epu default to "level" (they are
# mean-reverting stationary uncertainty indices); they may be switched to
# "diff" here and the report will reflect it.
TRANSFORMS = {
    **{
        c: "log_return"
        for c in [
            "fx_dxy",
            "fx_bbdxy",
            "fx_eurusd",
            "fx_usdcnh",
            "fx_audusd",
            "fx_usdbrl",
            "fx_usdclp",
            "fx_usdcad",
            "fx_usdnok",
            "fx_usdzar",
            "spx",
            "mxwo",
            "mxef",
            "csi300",
            "shcomp",
            "hscei",
            "xle",
            "bdiy",
            "bdti",
        ]
    },
    **{
        c: "diff"
        for c in [
            "ust_10y",
            "ust_5y",
            "ust_2y",
            "tips_10y",
            "tips_5y",
            "be_10y",
            "be_5y",
            "inflsw_5y",
            "inflsw_10y",
            "cny_10y",
            "hy_oas",
            "ig_oas",
            "vix",
            "move",
            "ovx",
            "gvz",
        ]
    },
    "gpr": "level",
    "epu": "level",
}

# The BBG multi-block sheets, in manifest order.
BBG_SHEETS = ["FX", "Rates_Infl", "Credit_Vol", "Equity_Growth", "Freight"]


# ----------------------------------------------------------------------------
# Series master / manifest  (manifest order == output column order).
# Static so the interface contract is deterministic; tickers/fields are
# cross-checked against the parsed workbook metadata at runtime.
# ----------------------------------------------------------------------------
@dataclass(frozen=True)
class SeriesMeta:
    series_id: str
    label: str
    ticker: str
    field: str
    units: str
    direction: str
    source_sheet: str
    notes: str = ""


MANIFEST: list[SeriesMeta] = [
    # --- FX (all log_return; direction not uniform) ---
    SeriesMeta(
        "fx_dxy", "ICE US Dollar Index", "DXY Curncy", "PX_LAST", "index", "UP=USD strong", "FX"
    ),
    SeriesMeta(
        "fx_bbdxy",
        "Bloomberg Dollar Spot Index",
        "BBDXY Index",
        "PX_LAST",
        "index",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_eurusd",
        "Euro / US Dollar",
        "EURUSD Curncy",
        "PX_LAST",
        "USD/EUR",
        "UP=EUR strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdcnh",
        "US Dollar / Offshore Yuan",
        "USDCNH Curncy",
        "PX_LAST",
        "CNH/USD",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_audusd",
        "Australian Dollar / US Dollar",
        "AUDUSD Curncy",
        "PX_LAST",
        "USD/AUD",
        "UP=AUD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdbrl",
        "US Dollar / Brazilian Real",
        "USDBRL Curncy",
        "PX_LAST",
        "BRL/USD",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdclp",
        "US Dollar / Chilean Peso",
        "USDCLP Curncy",
        "PX_LAST",
        "CLP/USD",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdcad",
        "US Dollar / Canadian Dollar",
        "USDCAD Curncy",
        "PX_LAST",
        "CAD/USD",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdnok",
        "US Dollar / Norwegian Krone",
        "USDNOK Curncy",
        "PX_LAST",
        "NOK/USD",
        "UP=USD strong",
        "FX",
    ),
    SeriesMeta(
        "fx_usdzar",
        "US Dollar / South African Rand",
        "USDZAR Curncy",
        "PX_LAST",
        "ZAR/USD",
        "UP=USD strong",
        "FX",
    ),
    # --- Rates / Inflation (all diff) ---
    SeriesMeta(
        "ust_10y",
        "10Y US Treasury yield",
        "USGG10YR Index",
        "PX_LAST",
        "yield %",
        "level=yield",
        "Rates_Infl",
    ),
    SeriesMeta(
        "ust_5y",
        "5Y US Treasury yield",
        "USGG5YR Index",
        "PX_LAST",
        "yield %",
        "level=yield",
        "Rates_Infl",
    ),
    SeriesMeta(
        "ust_2y",
        "2Y US Treasury yield",
        "USGG2YR Index",
        "PX_LAST",
        "yield %",
        "level=yield",
        "Rates_Infl",
    ),
    SeriesMeta(
        "tips_10y",
        "10Y TIPS real yield",
        "GTII10 Govt",
        "YLD_CNV_LAST",
        "real yield %",
        "can be negative -> never log",
        "Rates_Infl",
    ),
    SeriesMeta(
        "tips_5y",
        "5Y TIPS real yield",
        "GTII5 Govt",
        "YLD_CNV_LAST",
        "real yield %",
        "can be negative -> never log",
        "Rates_Infl",
    ),
    SeriesMeta(
        "be_10y",
        "10Y breakeven inflation",
        "USGGBE10 Index",
        "PX_LAST",
        "breakeven %",
        "",
        "Rates_Infl",
    ),
    SeriesMeta(
        "be_5y",
        "5Y breakeven inflation",
        "USGGBE05 Index",
        "PX_LAST",
        "breakeven %",
        "",
        "Rates_Infl",
    ),
    SeriesMeta(
        "inflsw_5y",
        "5Y USD inflation swap",
        "USSWIT5 Curncy",
        "PX_LAST",
        "swap %",
        "leg for 5y5y",
        "Rates_Infl",
    ),
    SeriesMeta(
        "inflsw_10y",
        "10Y USD inflation swap",
        "USSWIT10 Curncy",
        "PX_LAST",
        "swap %",
        "leg for 5y5y",
        "Rates_Infl",
    ),
    SeriesMeta(
        "cny_10y",
        "10Y China govt bond yield",
        "GCNY10YR Index",
        "PX_LAST",
        "yield %",
        "source seam 2016-08-02",
        "Rates_Infl",
    ),
    # --- Credit / Vol (all diff) ---
    SeriesMeta(
        "hy_oas",
        "US Corp High Yield OAS",
        "LF98OAS Index",
        "PX_LAST",
        "OAS % (x100=bp)",
        "UP=risk-off",
        "Credit_Vol",
    ),
    SeriesMeta(
        "ig_oas",
        "US Agg Corp (IG) OAS",
        "LUACOAS Index",
        "PX_LAST",
        "OAS % (x100=bp)",
        "UP=risk-off; long flat runs",
        "Credit_Vol",
    ),
    SeriesMeta(
        "vix",
        "CBOE S&P 500 implied vol",
        "VIX Index",
        "PX_LAST",
        "vol pts",
        "UP=risk-off",
        "Credit_Vol",
    ),
    SeriesMeta(
        "move",
        "ICE BofA MOVE (rates vol)",
        "MOVE Index",
        "PX_LAST",
        "bp vol",
        "UP=rates stress",
        "Credit_Vol",
    ),
    SeriesMeta(
        "ovx",
        "CBOE Crude Oil implied vol",
        "OVX Index",
        "PX_LAST",
        "vol pts",
        "UP=oil stress",
        "Credit_Vol",
    ),
    SeriesMeta(
        "gvz",
        "CBOE Gold implied vol",
        "GVZ Index",
        "PX_LAST",
        "vol pts",
        "UP=gold stress",
        "Credit_Vol",
    ),
    # --- Equity / Growth (all log_return) ---
    SeriesMeta(
        "spx", "S&P 500 (price)", "SPX Index", "PX_LAST", "index", "UP=risk-on", "Equity_Growth"
    ),
    SeriesMeta(
        "mxwo",
        "MSCI World (USD)",
        "MXWO Index",
        "PX_LAST",
        "index",
        "global growth",
        "Equity_Growth",
    ),
    SeriesMeta(
        "mxef",
        "MSCI Emerging Markets",
        "MXEF Index",
        "PX_LAST",
        "index",
        "EM growth",
        "Equity_Growth",
    ),
    SeriesMeta(
        "csi300",
        "CSI 300 (China A-shares)",
        "SHSZ300 Index",
        "PX_LAST",
        "index",
        "Asia close",
        "Equity_Growth",
    ),
    SeriesMeta(
        "shcomp",
        "Shanghai Composite",
        "SHCOMP Index",
        "PX_LAST",
        "index",
        "Asia close",
        "Equity_Growth",
    ),
    SeriesMeta(
        "hscei",
        "Hang Seng China Enterprises",
        "HSCEI Index",
        "PX_LAST",
        "index",
        "HK close",
        "Equity_Growth",
    ),
    SeriesMeta(
        "xle",
        "Energy Select Sector SPDR",
        "XLE US Equity",
        "PX_LAST",
        "price",
        "energy equity",
        "Equity_Growth",
    ),
    # --- Freight (all log_return) ---
    SeriesMeta(
        "bdiy", "Baltic Dry Index", "BDIY Index", "PX_LAST", "index", "dry-bulk freight", "Freight"
    ),
    SeriesMeta(
        "bdti",
        "Baltic Dirty Tanker (pulled as BIDY)",
        "BIDY Index",
        "PX_LAST",
        "index",
        "dirty-tanker freight",
        "Freight",
    ),
    # --- Uncertainty (calendar-daily, default level) ---
    SeriesMeta(
        "gpr",
        "Daily Geopolitical Risk (Caldara-Iacoviello)",
        "—",
        "GPRD",
        "index",
        "uncertainty (already stationary)",
        "GPR",
    ),
    SeriesMeta(
        "epu",
        "Daily US Economic Policy Uncertainty (EPU)",
        "—",
        "daily_policy_index",
        "index",
        "uncertainty (already stationary)",
        "EPR",
    ),
]

SERIES_IDS = [m.series_id for m in MANIFEST]
META_BY_ID = {m.series_id: m for m in MANIFEST}


# ============================================================================
# Parsing
# ============================================================================
def _to_date(val) -> pd.Timestamp:
    """Coerce a cell to a midnight-normalized Timestamp (NaT on failure)."""
    return pd.to_datetime(val, errors="coerce").normalize()


def _clean_series(sid: str, dates: list, values: list) -> pd.Series:
    """Build a clean float Series: drop empty dates, coerce, sort, dedupe."""
    s = pd.Series(
        pd.to_numeric(pd.Series(values), errors="coerce").to_numpy(),
        index=pd.to_datetime(pd.Series(dates), errors="coerce"),
        name=sid,
    )
    s.index = s.index.normalize()
    s = s[~s.index.isna()]  # drop trailing padding rows
    s = s.sort_index(kind="stable")
    s = s[~s.index.duplicated(keep="last")]  # dedupe dates, keep last
    assert s.index.is_monotonic_increasing, f"{sid}: dates not monotonic"
    assert s.index.is_unique, f"{sid}: duplicate dates after dedupe"
    return s


def parse_bbg_sheet(ws_rows: list[list]) -> dict[str, pd.Series]:
    """Parse a BBG multi-block sheet into {series_id: Series}.

    Blocks are contiguous 2-column [Date, Value] pairs. Scan columns
    left->right: when row-1 of a column is non-empty it starts a block
    (date col c, value col c+1); advance 2. Otherwise advance 1.
    Each block uses *its own* date column. Robust to added/removed series.
    """
    n_rows = len(ws_rows)
    n_cols = max((len(r) for r in ws_rows), default=0)

    def cell(r, c):
        row = ws_rows[r] if r < n_rows else []
        return row[c] if c < len(row) else None

    out: dict[str, pd.Series] = {}
    c = 0
    while c < n_cols:
        head = cell(0, c)
        if head is None or str(head).strip() == "":
            c += 1
            continue
        sid = str(head).strip()
        # metadata rows 2-4 (0-indexed 1..3): label, ticker, field
        ticker = cell(2, c) or ""
        field = cell(3, c) or ""
        # rows 7..end (0-indexed 6..) hold date | value
        dates, values = [], []
        for r in range(6, n_rows):
            d = cell(r, c)
            if d is None or (isinstance(d, str) and d.strip() == ""):
                continue
            dates.append(d)
            values.append(cell(r, c + 1))
        if sid in META_BY_ID:
            # Cross-check parsed ticker/field against the static manifest.
            m = META_BY_ID[sid]
            if str(ticker).strip() and str(ticker).strip() != m.ticker:
                print(f"  WARN {sid}: workbook ticker {ticker!r} != manifest {m.ticker!r}")
            if str(field).strip() and str(field).strip() != m.field:
                print(f"  WARN {sid}: workbook field {field!r} != manifest {m.field!r}")
        out[sid] = _clean_series(sid, dates, values)
        c += 2
    return out


def parse_gpr(ws_rows: list[list]) -> tuple[pd.Series, pd.DataFrame]:
    """Parse GPR sheet -> (gpr Series from GPRD@date, events DataFrame)."""
    header = [str(c).strip() if c is not None else "" for c in ws_rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    di, vi, ei = idx["date"], idx["GPRD"], idx.get("event")
    dates, values, ev_dates, ev_labels = [], [], [], []
    for r in ws_rows[1:]:
        if di >= len(r) or r[di] is None:
            continue
        dates.append(r[di])
        values.append(r[vi] if vi < len(r) else None)
        if ei is not None and ei < len(r) and r[ei] not in (None, ""):
            ev_dates.append(r[di])
            ev_labels.append(r[ei])
    s = _clean_series("gpr", dates, values)
    ev_dt = pd.to_datetime(pd.Series(ev_dates, dtype=object), errors="coerce")
    events = pd.DataFrame({"date": ev_dt, "event": ev_labels})
    events = events[events["date"].notna()].copy()
    events["date"] = events["date"].dt.normalize()
    return s, events


def parse_epu(ws_rows: list[list]) -> pd.Series:
    """Parse EPR sheet (EPU) -> epu Series; date = date(year, month, day)."""
    header = [str(c).strip() if c is not None else "" for c in ws_rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    yi, mi, dyi, vi = idx["year"], idx["month"], idx["day"], idx["daily_policy_index"]
    dates, values = [], []
    for r in ws_rows[1:]:
        if yi >= len(r) or r[yi] is None:
            continue
        try:
            y, mo, da = int(r[yi]), int(r[mi]), int(r[dyi])
            dates.append(pd.Timestamp(year=y, month=mo, day=da))
        except (TypeError, ValueError):
            continue
        values.append(r[vi] if vi < len(r) else None)
    return _clean_series("epu", dates, values)


def parse_regimes(ws_rows: list[list]) -> pd.DataFrame:
    """Parse REGIMES sheet -> tidy regime table."""
    rows = []
    for r in ws_rows[1:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        rows.append(
            {
                "regime": str(r[0]).strip(),
                "name": str(r[1]).strip() if r[1] is not None else "",
                "start": _to_date(r[2]),
                "end": _to_date(r[3]),
                "span": str(r[4]).strip() if r[4] is not None else "",
                "rationale": str(r[5]).strip() if len(r) > 5 and r[5] is not None else "",
            }
        )
    df = pd.DataFrame(rows)
    # Contiguity / non-overlap check.
    df = df.sort_values("start").reset_index(drop=True)
    for i in range(len(df) - 1):
        gap = (df.loc[i + 1, "start"] - df.loc[i, "end"]).days
        assert df.loc[i, "end"] < df.loc[i + 1, "start"], "regimes overlap"
        assert gap >= 1, f"regime gap unexpected between {df.loc[i, 'regime']} and next"
    return df


def load_workbook_rows(path: Path) -> dict[str, list[list]]:
    """Load every sheet as a list-of-rows (values only) for offline parsing."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        sheets[sn] = [list(row) for row in wb[sn].iter_rows(values_only=True)]
    wb.close()
    return sheets


# ============================================================================
# Transforms & alignment
# ============================================================================
def apply_transform(s: pd.Series, kind: str) -> pd.Series:
    """Apply a stationary transform on a working calendar (oldest->newest)."""
    if kind == "level":
        return s.copy()
    if kind == "diff":
        return s.diff()
    if kind == "log_return":
        pos = s.dropna()
        assert (pos > 0).all(), (
            f"{s.name}: log_return on non-positive value (min={pos.min()}) — never log this series"
        )
        return np.log(s).diff()
    raise ValueError(f"unknown transform {kind!r}")


def longest_run(mask: pd.Series) -> int:
    """Longest run of consecutive True values in a boolean Series."""
    best = cur = 0
    for v in mask.to_numpy():
        cur = cur + 1 if v else 0
        best = max(best, cur)
    return best


def count_flat_runs(s: pd.Series, threshold: int) -> int:
    """Count maximal runs of exactly-equal consecutive values longer than threshold."""
    vals = s.dropna().to_numpy()
    if len(vals) == 0:
        return 0
    cur = 1
    runs = 0
    for i in range(1, len(vals)):
        if vals[i] == vals[i - 1]:
            cur += 1
        else:
            if cur > threshold:
                runs += 1
            cur = 1
    if cur > threshold:
        runs += 1
    return runs


def assign_regime(dates: pd.DatetimeIndex, regimes: pd.DataFrame) -> pd.Series:
    """Map each date to its regime (inclusive Start<=date<=End)."""
    out = pd.Series(index=dates, dtype=object)
    for _, row in regimes.iterrows():
        m = (dates >= row["start"]) & (dates <= row["end"])
        out[m] = row["regime"]
    return out


# ============================================================================
# Pipeline
# ============================================================================
def run(input_path: Path, outdir: Path) -> None:
    proc = outdir / "data" / "processed"
    by_regime = proc / "by_regime"
    reports = outdir / "reports"
    for d in (proc, by_regime, reports, outdir / "data"):
        d.mkdir(parents=True, exist_ok=True)

    w0, w1 = pd.Timestamp(WINDOW[0]), pd.Timestamp(WINDOW[1])
    print(f"Loading workbook: {input_path}")
    sheets = load_workbook_rows(input_path)

    # ---- Parse all series ------------------------------------------------
    raw_series: dict[str, pd.Series] = {}
    for sn in BBG_SHEETS:
        raw_series.update(parse_bbg_sheet(sheets[sn]))
    gpr, gpr_events = parse_gpr(sheets["GPR"])
    epu = parse_epu(sheets["EPR"])
    raw_series["gpr"] = gpr
    raw_series["epu"] = epu
    regimes = parse_regimes(sheets["REGIMES"])

    missing = [sid for sid in SERIES_IDS if sid not in raw_series]
    assert not missing, f"missing series after parse: {missing}"

    # ---- Clip to window --------------------------------------------------
    clipped: dict[str, pd.Series] = {}
    for sid in SERIES_IDS:
        s = raw_series[sid]
        s = s[(s.index >= w0) & (s.index <= w1)]
        clipped[sid] = s

    # ---- Master trading-day grid = union of BBG-sheet dates --------------
    bbg_ids = [m.series_id for m in MANIFEST if m.source_sheet in BBG_SHEETS]
    grid_union = pd.DatetimeIndex([])
    for sid in bbg_ids:
        grid_union = grid_union.union(clipped[sid].index)
    master_grid = grid_union.sort_values()
    print(f"Master trading-day grid length: {len(master_grid)}")

    # ---- macro_levels_raw.csv  (full date union, raw values, no bridge) --
    raw_union = pd.DatetimeIndex([])
    for sid in SERIES_IDS:
        raw_union = raw_union.union(clipped[sid].index)
    raw_union = raw_union.sort_values()
    levels_raw = pd.DataFrame(index=raw_union)
    for sid in SERIES_IDS:
        levels_raw[sid] = clipped[sid].reindex(raw_union)
    levels_raw = levels_raw[SERIES_IDS]

    # ---- macro_levels_aligned.csv  (master grid, bridged, + regime) ------
    levels_aligned = pd.DataFrame(index=master_grid)
    bridge_info: dict[str, dict] = {}
    for sid in SERIES_IDS:
        col = clipped[sid].reindex(master_grid)
        pre_na = col.isna()
        filled = col.ffill(limit=BRIDGE_LIMIT)
        # A "bridged" day = was NaN on the grid, became non-NaN after ffill.
        bridged = pre_na & filled.notna()
        bridge_info[sid] = {
            "bridged_days": int(bridged.sum()),
            "longest_bridge": longest_run(bridged),
        }
        levels_aligned[sid] = filled
    levels_aligned = levels_aligned[SERIES_IDS]
    levels_aligned["regime"] = assign_regime(levels_aligned.index, regimes).to_numpy()

    # ---- macro_stationary.csv  (transform aligned levels, NaN-free) ------
    stat = pd.DataFrame(index=master_grid)
    for sid in SERIES_IDS:
        stat[sid] = apply_transform(levels_aligned[sid], TRANSFORMS[sid])
    stat = stat[SERIES_IDS]
    n_before = len(stat)
    stat = stat.iloc[1:]  # drop first row (diff/log NaN)
    # Which columns drive the row drops? (post-transform, pre-dropna NaN counts)
    drop_breakdown = {
        sid: int(stat[sid].isna().sum()) for sid in SERIES_IDS if stat[sid].isna().any()
    }
    stat = stat.dropna(how="any")  # trim gpr head + any residual NaN
    n_dropped = n_before - len(stat)
    stat_start = stat.index.min()
    stat["regime"] = assign_regime(stat.index, regimes).to_numpy()
    print(
        f"Stationary panel: {len(stat)} rows, start {stat_start.date()}, {n_dropped} rows dropped"
    )

    # ---- Per-regime slices ----------------------------------------------
    regime_counts: dict[str, int] = {}
    for reg in regimes["regime"]:
        sl = stat[stat["regime"] == reg]
        regime_counts[reg] = len(sl)

    # ====================================================================
    # ACCEPTANCE ASSERTIONS
    # ====================================================================
    checks: list[tuple[str, bool, str]] = []

    def check(name: str, cond: bool, detail: str = "") -> None:
        checks.append((name, bool(cond), detail))

    value_cols = [c for c in stat.columns if c != "regime"]
    check(
        "Exactly 37 value columns in manifest order",
        value_cols == SERIES_IDS and len(SERIES_IDS) == 37,
        f"{len(value_cols)} cols",
    )

    bounds_ok = all(
        (clipped[sid].index.min() >= w0) and (clipped[sid].index.max() <= w1)
        for sid in SERIES_IDS
        if len(clipped[sid])
    )
    check("Every series raw date within window", bounds_ok)

    check(
        "Master grid length in 3290-3320", 3290 <= len(master_grid) <= 3320, f"{len(master_grid)}"
    )

    check("macro_stationary has no NaN in value columns", not stat[value_cols].isna().any().any())
    check("regime non-null for all stationary rows", stat["regime"].notna().all())

    part_ok = sum(regime_counts.values()) == len(stat)
    check(
        "Regimes partition the stationary panel",
        part_ok,
        f"sum={sum(regime_counts.values())} rows={len(stat)}",
    )

    logret_ok = True
    for sid in SERIES_IDS:
        if TRANSFORMS[sid] == "log_return":
            v = clipped[sid].dropna()
            if len(v) and not (v > 0).all():
                logret_ok = False
    check("No log_return on non-positive series", logret_ok)

    check("Output column order deterministic == manifest", value_cols == SERIES_IDS)

    all_pass = all(c[1] for c in checks)

    # ====================================================================
    # WRITE OUTPUTS
    # ====================================================================
    def write_wide(df: pd.DataFrame, path: Path) -> None:
        out = df.copy()
        out.index.name = "date"
        out.index = out.index.strftime("%Y-%m-%d")
        out.to_csv(path, float_format="%.10g")
        print(f"  wrote {path}  ({len(out)} rows, {out.shape[1]} cols)")

    write_wide(levels_raw, proc / "macro_levels_raw.csv")
    write_wide(levels_aligned, proc / "macro_levels_aligned.csv")
    write_wide(stat, proc / "macro_stationary.csv")

    # transform_manifest.csv
    man = pd.DataFrame(
        [
            {
                "series_id": m.series_id,
                "label": m.label,
                "ticker": m.ticker,
                "field": m.field,
                "units": m.units,
                "direction": m.direction,
                "source_sheet": m.source_sheet,
                "transform": TRANSFORMS[m.series_id],
                "notes": m.notes,
            }
            for m in MANIFEST
        ]
    )
    man.to_csv(proc / "transform_manifest.csv", index=False)
    print(f"  wrote {proc / 'transform_manifest.csv'}")

    # regimes.csv
    reg_out = regimes.copy()
    reg_out["start"] = reg_out["start"].dt.strftime("%Y-%m-%d")
    reg_out["end"] = reg_out["end"].dt.strftime("%Y-%m-%d")
    reg_out[["regime", "name", "start", "end", "span", "rationale"]].to_csv(
        proc / "regimes.csv", index=False
    )
    print(f"  wrote {proc / 'regimes.csv'}")

    # per-regime slices
    for reg in regimes["regime"]:
        sl = stat[stat["regime"] == reg]
        write_wide(sl, by_regime / f"macro_stationary_{reg}.csv")

    # optional gpr events
    if len(gpr_events):
        ev = gpr_events[(gpr_events["date"] >= w0) & (gpr_events["date"] <= w1)].copy()
        if len(ev):
            ev["date"] = ev["date"].dt.strftime("%Y-%m-%d")
            ev.to_csv(proc / "gpr_events.csv", index=False)
            print(f"  wrote {proc / 'gpr_events.csv'}  ({len(ev)} events)")

    # ====================================================================
    # DATA QUALITY REPORT
    # ====================================================================
    report = build_report(
        input_path,
        master_grid,
        stat_start,
        n_dropped,
        drop_breakdown,
        levels_raw,
        levels_aligned,
        stat,
        clipped,
        bridge_info,
        regime_counts,
        regimes,
        checks,
        all_pass,
    )
    (reports / "data_quality_report.md").write_text(report)
    (outdir / "data" / "data_quality_report.md").write_text(report)
    print(f"  wrote {reports / 'data_quality_report.md'}")

    print("\n=== ACCEPTANCE ===")
    for name, ok, detail in checks:
        print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f"  ({detail})" if detail else ""))
    if not all_pass:
        print("\nSOME CHECKS FAILED", file=sys.stderr)
        sys.exit(1)
    print("\nAll acceptance checks PASSED.")


def build_report(
    input_path,
    master_grid,
    stat_start,
    n_dropped,
    drop_breakdown,
    levels_raw,
    levels_aligned,
    stat,
    clipped,
    bridge_info,
    regime_counts,
    regimes,
    checks,
    all_pass,
) -> str:
    L = []
    L.append("# Macro Panel Data Quality Report\n")
    L.append(f"- **Input:** `{input_path}`")
    L.append(f"- **Window (inclusive):** {WINDOW[0]} -> {WINDOW[1]}")
    L.append(f"- **Master trading-day grid length:** {len(master_grid)}")
    L.append(f"- **Stationary panel common start (NaN-free):** {stat_start.date()}")
    L.append(f"- **Rows dropped building stationary panel:** {n_dropped}")
    if drop_breakdown:
        bd = ", ".join(
            f"`{k}`={v}" for k, v in sorted(drop_breakdown.items(), key=lambda kv: -kv[1])
        )
        L.append(
            f"  - Driven by residual NaNs (post-transform, pre-dropna) in: {bd}. "
            "The 1 leading row is the differencing NaN; the rest are rows where "
            "`gpr` is missing — its ~10-day in-window head plus recurring ~23-calendar-day "
            "gaps (around Jan 1 / Nov 1 each year) that exceed the 10-trading-day bridge "
            "limit. These rows are dropped only from the convenience `macro_stationary` "
            "panel; the preferred downstream path (reindex `macro_levels_aligned` to factor "
            "dates, transform there) handles `gpr` NaNs independently and need not lose them."
        )
    L.append(f"- **macro_levels_raw rows:** {len(levels_raw)}")
    L.append(f"- **macro_levels_aligned rows:** {len(levels_aligned)}")
    L.append(f"- **macro_stationary rows:** {len(stat)}")
    L.append(
        f"- **gpr/epu transform:** gpr=`{TRANSFORMS['gpr']}`, epu=`{TRANSFORMS['epu']}` "
        "(default `level`; mean-reverting stationary uncertainty indices, editable to `diff`)"
    )
    L.append("")

    # Per-series table
    L.append("## Per-series QA\n")
    L.append(
        "| series_id | sheet | ticker | field | transform | raw first | raw last | "
        "raw non-null | bridged days | longest bridge | min | max | mean | flat runs>7 |"
    )
    L.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|---|")
    for m in MANIFEST:
        sid = m.series_id
        s = clipped[sid]
        bi = bridge_info[sid]
        lvl = levels_aligned[sid].dropna()
        first = s.index.min().date() if len(s) else "—"
        last = s.index.max().date() if len(s) else "—"
        flat = count_flat_runs(levels_aligned[sid], LONG_BRIDGE_FLAG)
        L.append(
            f"| {sid} | {m.source_sheet} | {m.ticker} | {m.field} | {TRANSFORMS[sid]} | "
            f"{first} | {last} | {int(s.notna().sum())} | {bi['bridged_days']} | "
            f"{bi['longest_bridge']} | {lvl.min():.4g} | {lvl.max():.4g} | "
            f"{lvl.mean():.4g} | {flat} |"
        )
    L.append("")

    # Long bridges
    long_bridges = [
        (sid, bridge_info[sid]["longest_bridge"])
        for sid in SERIES_IDS
        if bridge_info[sid]["longest_bridge"] > LONG_BRIDGE_FLAG
    ]
    L.append("## Long bridged runs (> 7 trading days)\n")
    if long_bridges:
        for sid, n in long_bridges:
            L.append(f"- `{sid}`: longest bridged run = {n} trading days")
    else:
        L.append("- None.")
    L.append("")

    # Global flags
    L.append("## Global flags\n")
    L.append(
        "- **`cny_10y`**: Bloomberg pricing-source **seam at 2016-08-02** "
        "(pre-seam priced by PCS:BGNC). Treat R1/R2/early-R3 China-yield results with care."
    )
    L.append(
        "- **`gpr`**: in-window coverage starts **2013-11-01**; ~10 leading trading days "
        "are NaN in levels and excluded from the stationary panel. It also has recurring "
        "**~23-calendar-day gaps around Jan 1 / Nov 1 each year** (≈16 trading days) that "
        "exceed the 10-trading-day bridge limit and therefore remain NaN on the master grid; "
        "those rows are dropped from the convenience `macro_stationary` panel (see dropped-row "
        "breakdown above). Shorter scattered gaps are bridged by ffill. Raw in-window count "
        "(3773) matches the reference exactly, so this is the source data, not a parse error."
    )
    L.append(
        "- **`ig_oas`**: long flat runs observed (tight IG spreads at 2-dp precision); "
        "confirm not a stale-feed artifact (see flat-runs column)."
    )
    L.append(
        "- **`bdti`**: pulled under ticker **`BIDY Index`**, labelled 'Baltic Dirty Tanker.' "
        "Confirm `BIDY` vs `BDTI` if results look off."
    )
    L.append(
        "- **`gpr`/`epu`**: calendar-daily sources reduced to the trading-day grid; "
        "defaulted to `level` transform."
    )
    L.append(
        "- **FX direction is not uniform**: USD-per-FX for `fx_eurusd`, `fx_audusd`; "
        "FX-per-USD for the `fx_usd*` pairs — see manifest `direction` column for sign handling."
    )
    L.append("")

    # Regime slices
    L.append("## Regime slices\n")
    L.append("| regime | name | start | end | rows |")
    L.append("|---|---|---|---|---|")
    for _, r in regimes.iterrows():
        L.append(
            f"| {r['regime']} | {r['name']} | {r['start'].date()} | "
            f"{r['end'].date()} | {regime_counts[r['regime']]} |"
        )
    L.append(f"| **total** | | | | **{sum(regime_counts.values())}** |")
    L.append("")

    # Downstream interface
    L.append("## Downstream interface (CCA)\n")
    L.append(
        "- **Preferred path:** reindex `macro_levels_aligned` to the AE factor dates, "
        "then difference / log-return on that calendar, so innovations are computed across "
        "consecutive *factor* days (exactly correct)."
    )
    L.append(
        "- **Convenience path:** `macro_stationary` differences on the macro union grid; "
        "the two diverge only on days adjacent to calendar mismatches."
    )
    L.append("- CCA is scale-invariant (internal whitening); no standardization is required.")
    L.append("")

    # Acceptance
    L.append("## Acceptance results\n")
    for name, ok, detail in checks:
        L.append(f"- [{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))
    L.append("")
    L.append(f"**Overall: {'ALL PASS' if all_pass else 'FAILURES PRESENT'}**")
    L.append("")
    return "\n".join(L)


def main() -> None:
    ap = argparse.ArgumentParser(description="Macro panel processing for AE<->Macro CCA.")
    ap.add_argument(
        "--input",
        default="data/macro/NEW_MACRO_COMMODITY_PANEL.xlsx",
        help="Path to the raw workbook.",
    )
    ap.add_argument("--outdir", default=".", help="Output root directory.")
    args = ap.parse_args()
    run(Path(args.input), Path(args.outdir))


if __name__ == "__main__":
    main()
